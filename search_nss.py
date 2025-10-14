"""
search_nss.py
Hybrid vyhledávání: xlsx data + Selenium pro NSS rozhodnutí
"""

import requests
import logging
import time
import random
from typing import List
from datetime import datetime
from pathlib import Path
import openpyxl

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

from models import Decision
from config import DATA_PATH

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# User-Agent rotace
USER_AGENTS = [
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Safari/605.1.15',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:120.0) Gecko/20100101 Firefox/120.0',
]


class NSSSearcher:
    """Hybrid vyhledávač pro NSS: xlsx metadata + Selenium pro texty"""

    XLSX_URL = "https://www.nssoud.cz/fileadmin/user_upload/dokumenty/Otevrena_data/Data_2025/Rijen/otevrena_data_NSS.xlsx"

    def __init__(self, delay: float = 2.0, use_selenium: bool = False):
        """
        Args:
            delay: Prodleva mezi požadavky (v sekundách)
            use_selenium: Použít Selenium pro získání textů
        """
        self.delay = delay
        self.use_selenium = use_selenium
        self.driver = None

        if use_selenium:
            self._init_selenium()

    def _init_selenium(self):
        """Inicializuje Selenium driver"""
        logger.info("🚀 Inicializuji Selenium...")

        options = Options()
        options.add_argument('--headless=new')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument(f'user-agent={random.choice(USER_AGENTS)}')

        service = Service(ChromeDriverManager().install())
        self.driver = webdriver.Chrome(service=service, options=options)
        logger.info("✅ Selenium připraven")

    def search_decisions(self, keywords: List[str], max_results: int = 100) -> List[Decision]:
        """
        Vyhledá rozhodnutí podle klíčových slov

        1. Stáhne xlsx s metadaty
        2. Vyfiltruje relevantní případy
        3. (Volitelně) Získá plné texty přes Selenium

        Args:
            keywords: Seznam klíčových slov
            max_results: Maximální počet výsledků

        Returns:
            Seznam objektů Decision
        """
        logger.info(f"🔍 Vyhledávám v NSS xlsx datech...")
        logger.info(f"📌 Klíčová slova: {', '.join(keywords)}")

        # 1. Stáhnout nebo načíst xlsx
        xlsx_path = self._download_xlsx()

        # 2. Načíst a vyfiltrovat data
        decisions = self._filter_xlsx_data(xlsx_path, keywords, max_results)

        logger.info(f"✅ Nalezeno {len(decisions)} rozhodnutí")

        # 3. (Volitelně) Získat texty přes Selenium
        if self.use_selenium and decisions:
            logger.info("📝 Získávám plné texty přes Selenium...")
            decisions = self._enrich_with_selenium(decisions)

        return decisions

    def _download_xlsx(self) -> Path:
        """Stáhne xlsx soubor s NSS daty"""
        xlsx_path = DATA_PATH / "nss_otevrena_data.xlsx"

        if xlsx_path.exists():
            # Kontrola stáří souboru (přestahnout pokud > 7 dní)
            age_days = (datetime.now().timestamp() - xlsx_path.stat().st_mtime) / 86400
            if age_days < 7:
                logger.info(f"📂 Používám cache: {xlsx_path} (stáří: {age_days:.1f} dní)")
                return xlsx_path

        logger.info(f"📥 Stahuji xlsx z: {self.XLSX_URL}")

        response = requests.get(self.XLSX_URL, timeout=60)
        response.raise_for_status()

        xlsx_path.write_bytes(response.content)
        logger.info(f"✅ Staženo: {len(response.content) / 1024 / 1024:.1f} MB")

        return xlsx_path

    def _filter_xlsx_data(self, xlsx_path: Path, keywords: List[str], max_results: int) -> List[Decision]:
        """Vyfiltruje relevantní rozhodnutí z xlsx"""
        logger.info("🔎 Filtruji xlsx data...")

        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        sheet = wb[wb.sheetnames[0]]

        # Najít sloupce
        headers = {cell.value: i for i, cell in enumerate(sheet[1], 1) if cell.value}

        decisions = []
        keyword_pattern = '|'.join(keywords).lower()

        for row_num, row in enumerate(sheet.iter_rows(min_row=2, max_row=min(100000, sheet.max_row)), 2):
            if len(decisions) >= max_results:
                break

            # Extrakce dat
            typ_veci = str(row[headers['Typ věci']-1].value or '')
            ucastnici = str(row[headers['Účastnící řízení s anonymizovanými fyzickými osobami']-1].value or '')
            spisova_znacka = str(row[headers['Spisová značka']-1].value or '')

            # Fulltextové hledání - hledá jednotlivá slova z klíčových slov
            search_text = f"{typ_veci} {ucastnici}".lower()

            # Rozdělit klíčová slova na jednotlivá slova
            keyword_words = []
            for kw in keywords:
                keyword_words.extend(kw.lower().split())

            if any(word in search_text for word in keyword_words):
                # Vytvoření Decision objektu
                decision = Decision(
                    ecli=f"CZ:NSS:{spisova_znacka.replace(' ', '-')}",
                    title=typ_veci or "Bez názvu",
                    date=self._parse_date(row[headers.get('Datum rozhodnutí', 'Došlo')-1].value),
                    url=f"https://vyhledavac.nssoud.cz/?spisova_znacka={spisova_znacka}",
                    keywords=keywords
                )

                # Přidat metadata
                decision.metadata = {
                    'spisova_znacka': spisova_znacka,
                    'soudce': str(row[headers.get('Soudce', 1)-1].value or ''),
                    'typ_rizeni': str(row[headers.get('Typ řízení', 1)-1].value or ''),
                    'typ_rozhodnuti': str(row[headers.get('Typ rozhodnutí', 1)-1].value or ''),
                }

                decisions.append(decision)

        wb.close()
        return decisions

    def _parse_date(self, date_value) -> datetime:
        """Parsuje datum z xlsx"""
        if isinstance(date_value, datetime):
            return date_value
        if isinstance(date_value, str):
            try:
                return datetime.strptime(date_value, "%Y-%m-%d")
            except:
                pass
        return None

    def _enrich_with_selenium(self, decisions: List[Decision]) -> List[Decision]:
        """Získá plné texty rozhodnutí pomocí Selenium"""
        enriched = []

        for i, decision in enumerate(decisions, 1):
            logger.info(f"📝 [{i}/{len(decisions)}] {decision.metadata.get('spisova_znacka')}")

            try:
                # Navigovat na detail rozhodnutí
                self.driver.get(decision.url)
                time.sleep(self.delay + random.uniform(-0.5, 0.5))  # Jitter

                # Hledat text rozhodnutí
                try:
                    # Různé možné selektory
                    text_element = WebDriverWait(self.driver, 10).until(
                        EC.presence_of_element_located((By.CLASS_NAME, "decision-text"))
                    )
                    decision.full_text = text_element.text
                    logger.info(f"  ✅ Text: {len(decision.full_text)} znaků")
                except:
                    logger.warning(f"  ⚠️  Text nenalezen")

                enriched.append(decision)

            except Exception as e:
                logger.error(f"  ❌ Chyba: {e}")
                enriched.append(decision)  # Přidat i bez textu

        return enriched

    def __del__(self):
        """Cleanup Selenium driveru"""
        if self.driver:
            self.driver.quit()


def search_decisions(keywords: List[str], max_results: int = 100, use_selenium: bool = False) -> List[Decision]:
    """
    Hlavní funkce pro vyhledávání rozhodnutí

    Args:
        keywords: Seznam klíčových slov
        max_results: Maximální počet výsledků
        use_selenium: Použít Selenium pro získání textů (pomalé!)

    Returns:
        Seznam objektů Decision
    """
    searcher = NSSSearcher(delay=2.0, use_selenium=use_selenium)
    return searcher.search_decisions(keywords, max_results)


if __name__ == "__main__":
    # Test
    test_keywords = ["územní plán"]
    results = search_decisions(test_keywords, max_results=5, use_selenium=False)

    print(f"\n📊 Nalezeno: {len(results)} rozhodnutí")
    for i, decision in enumerate(results, 1):
        print(f"\n{i}. {decision.title}")
        print(f"   ECLI: {decision.ecli}")
        print(f"   Datum: {decision.date}")
        print(f"   Spisová značka: {decision.metadata.get('spisova_znacka')}")
        print(f"   Soudce: {decision.metadata.get('soudce')}")
